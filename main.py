import openpyxl
import uuid


class Parser(object):
    file_name: str = ""
    brand_models: dict = {}
    sql_brands: str = ""
    sql_models: str = ""

    def __init__(self, fn: str):
        if self.file_name == "":
            self.file_name = fn

    def run(self):
        if self.file_name == "":
            print("invalid file name")
            return

        work_book = openpyxl.load_workbook(self.file_name)
        sheet = work_book.active

        counter: int = 1
        rows_count: int = sheet.max_row

        brand_name: str = ""

        while rows_count != counter - 1:
            brand: str = str(sheet.cell(row=counter, column=1).value).strip()
            model: str = str(sheet.cell(row=counter, column=2).value).strip()

            if brand_name != brand and brand != "" and brand != "None":
                brand_name = brand

            models = model.split("\n")

            if brand_name in self.brand_models:
                self.brand_models[brand_name].get("models").extend(models)
            else:
                self.brand_models[brand_name] = {
                    "brand_id": str(uuid.uuid4()),
                    "models": models
                }


            counter += 1

        return self

    def toSql(self):
        self.sql_brands = "INSERT INTO brands(id, title)"
        self.sql_models = "INSERT INTO models(id, brand_id, title)"

        for k in self.brand_models:
            self.sql_brands = (
                f"{self.sql_brands} "
                f"({dict(self.brand_models[k]).get('brand_id')}, '{k}'), \n"
            )

        for k in self.brand_models:
            for m in list(dict(self.brand_models[k]).get('models')):
                self.sql_models = (
                    f"{self.sql_models} "
                    f"({str(uuid.uuid4())}, "
                    f"{dict(self.brand_models[k]).get('brand_id')}, '{m}'), \n"
                )

        return self


parser = Parser(input("Enter file name: "))

parser.run().toSql()

print(parser.sql_brands)
print(parser.sql_models)